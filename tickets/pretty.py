# -*- coding: utf-8 -*-
"""
Reformat the ticket reports.

Input: XLSX file created by daily.py or weekly.py
Output: XLSX file reformatted.
"""
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import sys
TOP_BORDER = Border(top=Side(border_style='thin'))
LEFT_BORDER = Border(left=Side(border_style='thin'))
LEFT_TOP_BORDER = Border(top=Side(border_style='thin'),
                                  left=Side(border_style='thin'))

def trace(level, template, *args):
    if _args.verbose >= level:
        print(template.format(*args))


def one_sheet(infile):
    global workbook
    workbook = load_workbook(infile)
    ws = workbook.active
    trace(2, 'sheet: {}', ws.title)
    row2 = ws[2]
    for cell in row2:
        cell.alignment = Alignment(textRotation=90, horizontal='center')
    date_style = NamedStyle(name='datestyle',
                            number_format='YYYY-MM-DD')
    for cell in ws['A']:
        cell.style = date_style
        cell.alignment = Alignment(horizontal='center')
        lastrow = cell.row
    cell = ws.cell(row=lastrow + 1, column=1, value='Total')
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = TOP_BORDER
    ws.column_dimensions['A'].width = 12
    datetot = None
    for cell in ws[1]:
        if cell.value == 'totprice':
            totprice = cell.col_idx
        if cell.value == 'datetot':
            datetot = cell.col_idx
    for cell in ws[get_column_letter(totprice)][2:]:
        cell.border = LEFT_BORDER
    for cell in ws[get_column_letter(datetot)][2:]:
        cell.border = LEFT_BORDER

    trace(2, 'lastrow = {}, datetot = {}, totprice = {}',
          lastrow, datetot, totprice)

    # Compute total visitor counts
    for col in ws.iter_cols(min_col=2, max_col=totprice - 1):
        total = 0
        for c in col[4:]:
            if c.value:
                total += int(c.value)
        cell = ws.cell(row=lastrow + 1, column=c.col_idx, value=total)
        cell.border = TOP_BORDER

    # Compute total paid
    for col in ws.iter_cols(min_col=totprice, max_col=datetot):
        total = 0
        for c in col[3:]:
            c.number_format = '0.00'
            if c.value:
                total += float(c.value)
        cell = ws.cell(row=lastrow + 1, column=c.col_idx, value=total)
        cell.number_format = '£0.00'
        cell.border = TOP_BORDER
    cell = ws.cell(row=lastrow + 1, column=totprice)
    cell.border = LEFT_TOP_BORDER
    cell = ws.cell(row=lastrow + 1, column=datetot)
    cell.border = LEFT_TOP_BORDER

    workbook.save(_args.outfile)


def main():
    one_sheet()


def getargs():
    parser = argparse.ArgumentParser(
        description='''
        Reformat the ticket reports produced by daily.py or weekly.py.
        ''')
    parser.add_argument('infile', help='''
         The XLSX file that has been created by daily.py or weekly.py''')
    parser.add_argument('outfile',
                        help='''output XLSX file.
        ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    if sys.version_info.major < 3 or sys.version_info.minor < 6:
        raise ImportError('requires Python 3.6')
    _args = getargs()
    main()
    print('End prettifier.')
