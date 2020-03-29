# -*- coding: utf-8 -*-
"""
Reformat the ticket reports.

Input: Directory containing XLSX files created by daily.py and weekly.py
Output: XLSX file containing one tab per input file.
"""
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import os
import os.path
import re
import sys

TOP_BORDER = Border(top=Side(border_style='thin'))
LEFT_BORDER = Border(left=Side(border_style='thin'))
LEFT_TOP_BORDER = Border(top=Side(border_style='thin'),
                         left=Side(border_style='thin'))
CENTER = Alignment(horizontal='center')
DATE_STYLE = NamedStyle(name='datestyle', number_format='YYYY-MM-DD')

# Pattern to match filename like "tickets_2018-02_weekly_other.xlsx" and
# extract "weekly_other".
NAMEPAT = r'tickets_\d{4}-\d{2}_(.*)\.xlsx'
NAMEPATYEAR = r'tickets_\d{4}_(.*)\.xlsx'


def trace(level, template, *args):
    if _args.verbose >= level:
        print(template.format(*args))


def copy_sheet(oldwb, wb, tab):
    """

    :param oldwb: The source workbook to copy
    :param wb: The target workbook to contain the new worksheet.
    :param tab: The name to give the new sheet.
    :return: The new worksheet
    """
    ws = wb.create_sheet(tab)
    oldws = oldwb.active
    for row in oldws.iter_rows():
        for oldcell in row:
            trace(3, 'oldcell: {}', oldcell.coordinate)
            # skip merged cells
            if oldcell.value is None:
                continue
            ws.cell(row=oldcell.row, column=oldcell.col_idx,
                    value=oldcell.value)
    return ws


def one_sheet(ws):
    trace(2, 'sheet: {}, max_row: {}, max_column: {}',
          ws.title, ws.max_row, ws.max_column)
    datetot = totprice = None
    lastrow = len(ws['A'])
    for cell in ws[1]:
        if cell.value == 'totprice':
            totprice = cell.col_idx
        if cell.value in ('datetot', 'weektot'):
            datetot = cell.col_idx
    trace(2, 'lastrow = {}, datetot = {}, totprice = {}',
          lastrow, datetot, totprice)
    ws.cell(row=1, column=2).alignment = CENTER
    ws.cell(row=1, column=totprice).alignment = CENTER
    ws.merge_cells(start_row=1, start_column=2,
                   end_row=1, end_column=totprice - 1)
    ws.merge_cells(start_row=1, start_column=totprice,
                   end_row=1, end_column=datetot - 1)

    for cell in ws[2]:
        cell.alignment = Alignment(textRotation=90, horizontal='center')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions[get_column_letter(datetot)].width = 10
    for cell in ws['A']:
        cell.style = DATE_STYLE
        cell.alignment = CENTER
    for cell in ws['B']:
        cell.border = LEFT_BORDER
    for cell in ws[get_column_letter(totprice)]:
        cell.border = LEFT_BORDER
    for cell in ws[get_column_letter(datetot)]:
        cell.border = LEFT_BORDER

    cell = ws.cell(row=lastrow + 1, column=1, value='Total')
    cell.font = Font(bold=True)
    cell.alignment = CENTER
    cell.border = TOP_BORDER

    # Compute total visitor counts
    c = None
    for col in ws.iter_cols(min_col=2, max_col=totprice - 1):
        total = 0
        for c in col[3:]:
            if c.value:
                total += int(c.value)
        cell = ws.cell(row=lastrow + 1, column=c.col_idx, value=total)
        cell.border = TOP_BORDER

    # Compute total paid
    for col in ws.iter_cols(min_col=totprice, max_col=datetot):
        total = 0
        for c in col[3:]:
            c.number_format = '0.00'
            if c.value:
                total += float(c.value)
        cell = ws.cell(row=lastrow + 1, column=c.col_idx, value=total)
        cell.number_format = '£0.00'
        cell.border = TOP_BORDER
    ws.cell(row=lastrow + 1, column=2).border = LEFT_TOP_BORDER
    ws.cell(row=lastrow + 1, column=totprice).border = LEFT_TOP_BORDER
    ws.cell(row=lastrow + 1, column=datetot).border = LEFT_TOP_BORDER


def main():
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]  # remove the default sheet
    for f in sorted(os.listdir(_args.indir)):
        trace(2, 'input: {}', f)
        if not f.endswith('.xlsx'):
            print(f'Unknown file "{f}" skipped. Not ending in .xlsx.')
            continue
        mat = re.match(NAMEPAT, f)
        if not mat:  # Try year pattern
            mat = re.match(NAMEPATYEAR, f)
        if not mat:
            print(f'Unknown file "{f}" skipped. Failed pattern match.')
            continue
        tabname = mat.group(1)
        if tabname == 'merged':  # in case we re-run the script
            trace(1, 'skipping: {}', f)
            continue
        trace(1, 'reading: {}', f)
        wbpath = os.path.join(_args.indir, f)
        oldworkbook = load_workbook(wbpath)
        ws = copy_sheet(oldworkbook, workbook, tabname)
        one_sheet(ws)
    print(f'Writing to: {_args.outfile}')
    workbook.save(_args.outfile)


def getargs():
    parser = argparse.ArgumentParser(
        description='''
        Reformat the ticket reports produced by daily.py or weekly.py.
        ''')
    parser.add_argument('indir', help='''
         The directory containing the XLSX files that have been created by
         daily.py and weekly.py''')
    parser.add_argument('outfile',
                        help='''output XLSX file.
        ''')
    parser.add_argument('-v', '--verbose', default=1, type=int, help='''
    Modify verbosity.
    ''')
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    assert sys.version_info >= (3, 6)
    _args = getargs()
    if _args.verbose > 1:
        print(f'verbosity: {_args.verbose}')
    main()
    print('End pretty2.')
